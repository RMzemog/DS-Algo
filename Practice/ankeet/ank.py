import openpyxl
from tkinter import *

wb=openpyxl.load_workbook("Book1.xlsx")
sheets=wb.sheetnames
print(sheets)
print(wb.active.title)
for j in range(len(sheets)):
    x = sheets[j]
    sh1=wb[x]
    for i in range(2,sh1.max_row+1):
        x=''
        for j in range(1,sh1.max_column+1):
            if j==2 or j in range(4,sh1.max_column):
                if sh1.cell(i,j).value==None:
                    x+="    "
                else:
                    x+="    " +str(sh1.cell(i,j).value)

        print(x)
root = Tk() 
frame = Frame(root) 
frame.pack() 
bottomframe = Frame(root) 
bottomframe.pack( side = BOTTOM ) 
redbutton = Button(frame, text = 'Red', fg ='red') 
redbutton.pack( side = LEFT) 
greenbutton = Button(frame, text = 'Brown', fg='brown') 
greenbutton.pack( side = LEFT ) 
bluebutton = Button(frame, text ='Blue', fg ='blue') 
bluebutton.pack( side = LEFT ) 
blackbutton = Button(bottomframe, text ='Black', fg ='black') 
blackbutton.pack( side = BOTTOM) 
root.mainloop() 